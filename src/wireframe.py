"""
wireframe.py — Campaign wireframe export, matching the Namshi 'June Onsite Wireframes' format.

One Excel sheet per campaign. Structure per sheet:
  - Header block  (rows 1-10): Campaign metadata
  - Per-module blocks: module label → EN copy → AR copy → IMG → Link rows
"""

from pathlib import Path
import calendar as cal_mod

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE       = Path(__file__).resolve().parent.parent
TILES_FILE = BASE / "config" / "wireframe_tiles.yaml"

# ── Colours matching the Namshi wireframe template ───────────────────────────
YELLOW_FILL  = PatternFill("solid", fgColor="FFFFFF00")   # "Onsite Plan" header
ORANGE_FILL  = PatternFill("solid", fgColor="FFFFD966")   # module header rows
GREY_FILL    = PatternFill("solid", fgColor="FFF2F2F2")   # sub-label rows (IMG / Link)
RED_TEXT     = Font(name="Arial", size=10, color="FFFF0000", bold=False)
BLUE_TEXT    = Font(name="Arial", size=10, color="FF1155CC", bold=False)
BOLD         = Font(name="Arial", size=10, bold=True)
NORMAL       = Font(name="Arial", size=10)
HEADER_FONT  = Font(name="Arial", size=11, bold=True, color="FFFF0000")   # "Onsite Plan"
MODULE_FONT  = Font(name="Arial", size=11, bold=True)
AR_FONT      = Font(name="Arial", size=10)

THIN  = Side(style="thin",  color="FFD0D0D0")
THICK = Side(style="medium", color="FF888888")
BORDER_THIN  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_THICK = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

# Column widths (A=labels, B-E=metadata values, F=sub-labels, G+=tile content)
COL_A_W = 22    # field labels
COL_B_W = 30    # metadata values
COL_F_W = 8     # sub-labels: IMG / Link / SKU
COL_G_W = 50    # first tile content column


def load_tile_catalog() -> list[dict]:
    if not TILES_FILE.exists():
        return []
    with open(TILES_FILE, encoding="utf-8") as f:
        data = yaml.safe_load(f) or {}
    return data.get("tiles", [])


def tile_catalog_map() -> dict:
    return {t["id"]: t for t in load_tile_catalog()}


def build_wireframe_summary(tiles) -> str:
    catalog = tile_catalog_map()
    parts = []
    for i, entry in enumerate(tiles, start=1):
        tid = entry.get("tile_id", entry) if isinstance(entry, dict) else entry
        label = catalog.get(tid, {}).get("label", tid)
        parts.append(f"{i}. {label}")
    return "  ".join(parts) if parts else "TBD"


# ── Cell writing helpers ─────────────────────────────────────────────────────

def _write(ws, row, col, value, font=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:  cell.font  = font
    if fill:  cell.fill  = fill
    if align: cell.alignment = align
    return cell


def _label_value(ws, row, label, value, value_font=None):
    """Write a label in col A and a value in cols B–E (merged)."""
    _write(ws, row, 1, label, font=BOLD,
           align=Alignment(horizontal="left", vertical="center"))
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    _write(ws, row, 2, value, font=value_font or BOLD,
           align=Alignment(horizontal="left", vertical="center", wrap_text=True))
    ws.row_dimensions[row].height = 16


def _section_header(ws, row, text, last_col):
    """Full-width orange module header row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = _write(ws, row, 1, text, font=MODULE_FONT, fill=ORANGE_FILL,
                  align=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[row].height = 20
    return cell


def _sub_row(ws, row, label, content, content_font=None, fill=None, last_col=10):
    """IMG / Link / Copy sub-row: label in col F, content in cols G–last_col merged."""
    _write(ws, row, 6, label, font=BOLD, fill=fill or GREY_FILL,
           align=Alignment(horizontal="center", vertical="center"))
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=last_col)
    _write(ws, row, 7, content, font=content_font or NORMAL, fill=fill,
           align=Alignment(horizontal="left", vertical="top", wrap_text=True))
    ws.row_dimensions[row].height = 14


# ── Main export function ─────────────────────────────────────────────────────

def write_wireframe_xlsx(campaigns: list[dict], filepath, month: int, year: int):
    """
    Write one Excel workbook with one sheet per campaign, matching the
    Namshi onsite wireframe template format.
    """
    catalog  = tile_catalog_map()
    wb       = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    for camp in campaigns:
        name       = camp.get("campaign_name") or "Untitled"
        wm         = camp.get("wireframe_meta") or {}
        tiles_raw  = camp.get("wireframe_tiles") or []

        # Normalise tile entries
        tiles = []
        for t in tiles_raw:
            if isinstance(t, str):
                tiles.append({"tile_id": t, "copy_en": "", "copy_en_refined": "",
                               "copy_ar": "", "cta_text": "", "deeplink": "",
                               "asset_dimensions": ""})
            else:
                tiles.append(t)

        # Sheet name: max 31 chars, no invalid chars
        safe = name[:31].replace("/", "-").replace("\\", "-").replace("*", "").replace(
            "?", "").replace("[", "").replace("]", "").replace(":", "-")
        ws = wb.create_sheet(title=safe)

        LAST_COL = 10  # A–J

        # Column widths
        ws.column_dimensions["A"].width = COL_A_W
        for col_letter in ["B", "C", "D", "E"]:
            ws.column_dimensions[col_letter].width = COL_B_W
        ws.column_dimensions["F"].width = COL_F_W
        ws.column_dimensions["G"].width = COL_G_W
        for col_letter in ["H", "I", "J"]:
            ws.column_dimensions[col_letter].width = 30

        # ── Header block ──────────────────────────────────────────────────
        row = 1

        # Row 1: "Onsite Plan"
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=LAST_COL)
        _write(ws, row, 1, "Onsite Plan", font=HEADER_FONT, fill=YELLOW_FILL,
               align=Alignment(horizontal="left", vertical="center", indent=1))
        ws.row_dimensions[row].height = 20
        row += 1

        # Row 2: Campaign Name
        _label_value(ws, row, "Project / Campaign Name", name)
        row += 1

        # Row 3: Overview (feeds + surfaces)
        overview_parts = []
        if wm.get("feeds"):    overview_parts.append(", ".join(wm["feeds"]) + " feed")
        if wm.get("surfaces"): overview_parts.append(", ".join(wm["surfaces"]))
        if not overview_parts: overview_parts.append(camp.get("category_focus") or "Fashion feed")
        _label_value(ws, row, "Overview", " + ".join(overview_parts))
        row += 1

        # Row 4: Placement on App
        surfaces_val = ", ".join(wm.get("surfaces") or []) or camp.get("geography") or "All"
        _label_value(ws, row, "Placement on App", surfaces_val)
        row += 1

        # Row 5: Market(s)
        markets_val = ", ".join(wm.get("countries") or []) or camp.get("geography") or "All"
        _label_value(ws, row, "Market(s)", markets_val)
        row += 1

        # Row 6: Start Date & Time
        _label_value(ws, row, "Start Date & Time",
                     wm.get("live_start") or camp.get("start_date") or "TBD")
        row += 1

        # Row 7: End Date & Time
        _label_value(ws, row, "End Date & Time",
                     wm.get("live_end") or camp.get("end_date") or "TBD")
        row += 1

        # Row 8: Deliverables (tile names)
        tile_labels = []
        for t in tiles:
            tid   = t.get("tile_id", "")
            label = catalog.get(tid, {}).get("label", tid)
            tile_labels.append(label)
        deliverables = " + ".join(tile_labels) if tile_labels else "TBD"
        _label_value(ws, row, "Deliverables", deliverables)
        row += 1

        # Row 9: Platform / Technical
        tech_parts = []
        if wm.get("platforms"):      tech_parts.append("Platform: " + ", ".join(wm["platforms"]))
        if wm.get("app_version_min"): tech_parts.append("Min ver: " + wm["app_version_min"])
        if wm.get("ab_test"):
            ab = "A/B: Yes"
            if wm.get("ab_variant"): ab += f" ({wm['ab_variant']})"
            tech_parts.append(ab)
        if tech_parts:
            _label_value(ws, row, "Technical", "  ·  ".join(tech_parts))
            row += 1

        # Row: Priority + Asset position
        ops_parts = []
        if wm.get("priority"):       ops_parts.append("Priority: " + ", ".join(wm["priority"]))
        if wm.get("asset_position"): ops_parts.append("Position: " + wm["asset_position"])
        if ops_parts:
            _label_value(ws, row, "Placement Details", "  ·  ".join(ops_parts))
            row += 1

        # Rows: Copy Required / Design Required
        any_copy = any(t.get("copy_en") or t.get("copy_en_refined") for t in tiles)
        _label_value(ws, row, "Copy Required?", "YES" if any_copy else "TBD")
        row += 1
        _label_value(ws, row, "Design Required?", "YES")
        row += 1

        # Notes
        if wm.get("notes"):
            _label_value(ws, row, "Notes / Brief", wm["notes"], value_font=NORMAL)
            ws.row_dimensions[row].height = max(
                16, min(80, 16 * (wm["notes"].count("\n") + 1)))
            row += 1

        # Spacer
        ws.row_dimensions[row].height = 8
        row += 1

        # ── Module blocks ─────────────────────────────────────────────────
        for idx, t in enumerate(tiles, start=1):
            tid   = t.get("tile_id", "")
            label = catalog.get(tid, {}).get("label", tid)

            copy_en   = (t.get("copy_en_refined") or t.get("copy_en") or "").strip()
            copy_ar   = (t.get("copy_ar") or "").strip()
            cta       = (t.get("cta_text") or "").strip()
            link      = (t.get("deeplink") or "").strip()
            dims      = (t.get("asset_dimensions") or "").strip()

            # Module header row (orange)
            header_text = f"{idx}. {label}"
            if dims: header_text += f"  |  {dims}"
            _section_header(ws, row, header_text, LAST_COL)
            row += 1

            # EN copy row
            en_lines = []
            if copy_en: en_lines.append(f"EN:  {copy_en}")
            if cta:     en_lines.append(f"CTA: {cta}")
            if not en_lines: en_lines.append("EN: [Copy TBD]")
            _sub_row(ws, row, "EN", "\n".join(en_lines),
                     content_font=BOLD, fill=None, last_col=LAST_COL)
            ws.row_dimensions[row].height = max(16, 16 * len(en_lines))
            row += 1

            # AR copy row
            if copy_ar:
                ar_cell_font = Font(name="Arial", size=10)
                _sub_row(ws, row, "AR", copy_ar,
                         content_font=ar_cell_font, fill=None, last_col=LAST_COL)
                # RTL alignment on the AR content cell
                ws.cell(row=row, column=7).alignment = Alignment(
                    horizontal="right", vertical="top", wrap_text=True, readingOrder=2)
                ws.row_dimensions[row].height = max(16, 16 * (copy_ar.count("\n") + 1))
                row += 1

            # IMG / SKU row
            _sub_row(ws, row, "IMG", "[Insert product SKU image URL]",
                     content_font=BLUE_TEXT, fill=GREY_FILL, last_col=LAST_COL)
            row += 1

            # Link row
            link_val = link if link else "[Insert landing page URL]"
            _sub_row(ws, row, "Link", link_val,
                     content_font=RED_TEXT if link else NORMAL,
                     fill=GREY_FILL, last_col=LAST_COL)
            row += 1

            # Spacer
            ws.row_dimensions[row].height = 6
            row += 1

        # Freeze the header rows
        ws.freeze_panes = ws.cell(row=13, column=1)

    if not wb.sheetnames:
        wb.create_sheet("No Campaigns")

    wb.save(filepath)
