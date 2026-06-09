"""
generate.py — Steps 3-5: Resolve target month's events → draft calendar XLSX + notes

Usage:
    python src/generate.py --month 7 --year 2026

Outputs:
    output/draft_calendar_<month>_<year>.xlsx
    output/draft_notes.md
"""

import argparse
import calendar
import sys
import logging
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
import openpyxl
import yaml
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE     = Path(__file__).resolve().parent.parent
CFG_FILE = BASE / "config" / "calendar_config.yaml"
CSV_FILE = BASE / "data" / "standardized_history.csv"
OUT_DIR  = BASE / "output"

sys.path.insert(0, str(BASE / "src"))
from recurrence import build_kb
from signals    import enrich_with_commercial_signals, LAYER2_COLS

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ── Schema ─────────────────────────────────────────────────────────────────────
SCHEMA_COLS = [
    "campaign_name", "canonical_key", "tier", "geography",
    "start_date", "end_date", "duration_days",
    "promos_coupons", "category_focus", "onsite_deliverables",
    "wireframes", "event_type", "source", "needs_review", "raw_notes",
] + LAYER2_COLS

# Display headers (human-friendly) — same order as SCHEMA_COLS
DISPLAY_HEADERS = [
    "Campaign Name", "Canonical Key", "Tier", "Geography",
    "Tentative Start Date", "Tentative End Date", "Duration (days)",
    "Promos / Coupons", "Category Focus", "Onsite Deliverables",
    "Wireframes", "Event Type", "Source", "Needs Review?", "Raw Notes",
    "Trending Keywords", "Stock Flag", "Brand Performance Flag",
    "Struggling Category Flag", "New Activation Flag", "Outlier Flag",
]

# Styling (matching june2026 template)
HEADER_FILL  = PatternFill("solid", fgColor="FF000000")   # black
HEADER_FONT  = Font(name="Calibri", bold=True, color="FF00FF00")  # green
DATA_FONT    = Font(name="Calibri")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
THIN         = Side(style="thin", color="FFCCCCCC")
CELL_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Approximate column widths
COL_WIDTHS = {
    "A": 35, "B": 18, "C": 12, "D": 14,
    "E": 20, "F": 20, "G": 14, "H": 30,
    "I": 35, "J": 45, "K": 25, "L": 16,
    "M": 10, "N": 14, "O": 30,
    "P": 25, "Q": 14, "R": 22, "S": 22, "T": 22, "U": 14,
}

LAYER2_PLACEHOLDER = "[Layer 2 — pending]"


# ── Config helpers ──────────────────────────────────────────────────────────────
def load_config():
    with open(CFG_FILE, encoding="utf-8") as f:
        return yaml.safe_load(f)


def month_bounds(year, month):
    _, last = calendar.monthrange(year, month)
    return date(year, month, 1), date(year, month, last)


def overlaps_month(start_str, end_str, year, month):
    """True if the date range overlaps with the target year/month."""
    m_start, m_end = month_bounds(year, month)
    try:
        s = date.fromisoformat(start_str)
        e = date.fromisoformat(end_str)
        return s <= m_end and e >= m_start
    except Exception:
        return False


def nth_weekday(year, month, weekday, n):
    """Return the nth occurrence of weekday (Mon=0 … Sun=6) in year/month."""
    count = 0
    for day in range(1, calendar.monthrange(year, month)[1] + 1):
        if date(year, month, day).weekday() == weekday:
            count += 1
            if count == n:
                return date(year, month, day)
    raise ValueError(f"No {n}th weekday={weekday} in {year}/{month}")


# ── Date estimation ─────────────────────────────────────────────────────────────
def estimate_dates(ev, kb_entry, year, month):
    """
    Estimate (start_iso, end_iso, is_estimated) for an event in the target month.
    Priority:
      1. Explicit config dates (_resolved_start / _resolved_end already set)
      2. Rule-based hard-coded anchors (Payday ~25th, Feed Refresh ~7th, etc.)
      3. History: most_recent_start/end from KB, offset +1yr if prior year
      4. History: typical_start_day + duration, but only if start_day fits target month
      5. Seasonal window fallback (full target month)
    """
    ck = ev["canonical_key"]

    # 1. Already resolved (lunar / sports events)
    if ev.get("_resolved_start"):
        return ev["_resolved_start"], ev.get("_resolved_end", "TBD"), False

    # 2. Rule-based hard-coded anchors  ← checked FIRST so they are never estimated
    if ck == "payday":
        start = date(year, month, 25)
        end   = start + timedelta(days=8)
        return start.isoformat(), end.isoformat(), False

    if ck == "feed_refresh":
        start = date(year, month, 7)
        return start.isoformat(), "TBD", False

    if ck == "fathers_day" and month == 6:
        d = nth_weekday(year, month, 6, 3)   # 3rd Sunday
        return d.isoformat(), d.isoformat(), False

    if "day" in ev:
        try:
            d = date(year, month, ev["day"])
            return d.isoformat(), d.isoformat(), False
        except Exception:
            pass

    if "rule" in ev and "third sunday" in ev["rule"].lower() and month == 6:
        d = nth_weekday(year, month, 6, 3)
        return d.isoformat(), d.isoformat(), False

    # 3. History: use most_recent_start/end from KB, offset to target year if needed
    if kb_entry:
        mr_start = kb_entry.get("most_recent_start")
        mr_end   = kb_entry.get("most_recent_end")
        if mr_start:
            try:
                s_date = date.fromisoformat(mr_start)
                e_date = date.fromisoformat(mr_end) if mr_end else None
                # Offset to target year if the history instance is from a prior year
                delta_yr = year - s_date.year
                s_adj    = s_date.replace(year=s_date.year + delta_yr)
                e_adj    = e_date.replace(year=e_date.year + delta_yr) if e_date else None
                # Use if the adjusted event overlaps the target month
                m_start_d, m_end_d = month_bounds(year, month)
                if s_adj <= m_end_d and (e_adj is None or e_adj >= m_start_d):
                    return (
                        s_adj.isoformat(),
                        e_adj.isoformat() if e_adj else "TBD",
                        bool(delta_yr > 0),   # is_estimated only if we offset by a year
                    )
            except Exception:
                pass

    # 4. History: typical_start_day anchored to THIS month (only if plausible)
    if kb_entry and kb_entry.get("typical_start_day"):
        typ_months = kb_entry.get("typical_months", [])
        # Only apply if this campaign typically starts in the target month
        if month in typ_months or not typ_months:
            sd = kb_entry["typical_start_day"]
            try:
                start = date(year, month, sd)
                dur   = kb_entry.get("typical_duration_days")
                end   = (start + timedelta(days=int(dur) - 1)).isoformat() if dur else "TBD"
                return start.isoformat(), end, True
            except Exception:
                pass

    # 5. Seasonal window fallback
    if "start_month" in ev:
        m_start, m_end = month_bounds(year, month)
        return m_start.isoformat(), m_end.isoformat(), True

    return "TBD", "TBD", True


def calc_duration(start, end):
    if start != "TBD" and end != "TBD":
        try:
            return (date.fromisoformat(end) - date.fromisoformat(start)).days + 1
        except Exception:
            pass
    return None


# ── Event resolution ────────────────────────────────────────────────────────────
def resolve_events(cfg, year, month):
    """
    Return list of event dicts (augmented with _resolved_start/_resolved_end
    for lunar/sports) that are active in year/month.
    """
    events    = []
    m_start, m_end = month_bounds(year, month)

    # 1. Fixed-date events
    for ev in cfg.get("fixed_date_events", []):
        ev_month = ev.get("month")
        # DSF spans Dec-Jan
        if ev_month == month or (ev.get("canonical_key") == "dsf" and month == 1):
            events.append(dict(ev))

    # 2. Seasonal windows
    for ev in cfg.get("seasonal_windows", []):
        sm, em = ev.get("start_month", 0), ev.get("end_month", 0)
        in_window = (sm <= month <= em) if sm <= em else (month >= sm or month <= em)
        # DSF-style winter window handled above; skip duplicate
        if in_window:
            events.append(dict(ev))

    # 3. Monthly anchors — always include Payday + Feed Refresh
    for ev in cfg.get("monthly_anchors", []):
        e = dict(ev)
        if ev["canonical_key"] == "beauty_week":
            e["_needs_review_override"] = True  # confirm active
        events.append(e)

    # 4. Lunar events — check per-year dates
    for ev in cfg.get("lunar_events", []):
        dates = ev.get("dates", {})
        yd    = dates.get(year) or dates.get(str(year))
        if yd:
            s, e_end = yd.get("start"), yd.get("end")
            if s and e_end and overlaps_month(s, e_end, year, month):
                entry = dict(ev)
                entry["_resolved_start"] = s
                entry["_resolved_end"]   = e_end
                events.append(entry)

    # 5. Sports events — check per-year dates
    for ev in cfg.get("sports_events", []):
        dates = ev.get("dates", {})
        yd    = dates.get(year) or dates.get(str(year))
        if yd:
            s, e_end = yd.get("start"), yd.get("end")
            if s and e_end and overlaps_month(s, e_end, year, month):
                entry = dict(ev)
                entry["_resolved_start"] = s
                entry["_resolved_end"]   = e_end
                events.append(entry)

    return events


# ── Row generation ──────────────────────────────────────────────────────────────
def build_row(ev, kb, year, month):
    """Build one draft row from an event + knowledge base entry."""
    ck       = ev["canonical_key"]
    kb_entry = kb.get(ck, {})

    start, end, is_est = estimate_dates(ev, kb_entry, year, month)
    needs_review = (
        ev.get("needs_date_confirmation", False)
        or ev.get("_needs_review_override", False)
        or is_est
        or not kb_entry
    )

    tier  = kb_entry.get("most_recent_tier")  or ev.get("default_tier", "-")
    geo   = kb_entry.get("most_recent_geography") or ev.get("geography", "KSA + UAE")
    cat   = kb_entry.get("most_recent_category_focus") or ev.get("category_focus", "")
    deliv = kb_entry.get("most_recent_onsite_deliverables") or ev.get("onsite_deliverables", "")

    return {
        "campaign_name":      ev["name"],
        "canonical_key":      ck,
        "tier":               tier,
        "geography":          geo,
        "start_date":         start,
        "end_date":           end,
        "duration_days":      calc_duration(start, end),
        "promos_coupons":     "TBC",
        "category_focus":     cat,
        "onsite_deliverables": deliv,
        "wireframes":         "TBD",
        "event_type":         ev.get("event_type", ""),
        "source":             "both" if kb_entry else "config",
        "needs_review":       needs_review,
        "raw_notes":          (
            f"Date confirmation required per config." if ev.get("needs_date_confirmation") else
            f"Dates estimated from history (offset +1yr)." if is_est and kb_entry else
            f"No historical precedent — config defaults used." if not kb_entry else ""
        ),
    }


def build_bts_rows(ev, kb, year, month):
    """
    BTS gets one row per phase active in target month.
    Phases: Soft Launch (mid-month), Takeover KSA, Takeover UAE.
    """
    kb_entry = kb.get("bts", {})
    rows     = []
    phases   = ev.get("phased", [])
    m_start  = month_bounds(year, month)[0]

    # July is start of BTS window → all three phases begin
    if month == 7:
        phase_defs = [
            ("BTS Soft Launch",    date(year, 7, 15), date(year, 7, 24), "KSA + UAE", "Tier 2"),
            ("BTS Takeover (KSA)", date(year, 7, 25), date(year, 8, 12), "KSA",       "Tier 2"),
            ("BTS Takeover (UAE)", date(year, 7, 25), date(year, 8, 20), "UAE",       "Tier 2"),
        ]
    elif month == 8:
        phase_defs = [
            ("BTS Takeover (KSA)", date(year, 7, 25), date(year, 8, 12), "KSA",       "Tier 2"),
            ("BTS Takeover (UAE)", date(year, 7, 25), date(year, 8, 20), "UAE",       "Tier 2"),
            ("BTS Last Chance",    date(year, 8, 20), date(year, 9,  4), "KSA + UAE", "Tier 2"),
        ]
    elif month == 9:
        phase_defs = [
            ("BTS Last Chance", date(year, 8, 25), date(year, 9, 4), "KSA + UAE", "Tier 2"),
        ]
    else:
        # Generic fallback
        phase_defs = [
            ("Back to School", m_start, m_start + timedelta(days=20), "KSA + UAE", "Tier 2"),
        ]

    cat   = kb_entry.get("most_recent_category_focus") or ev.get("category_focus", "")
    deliv = kb_entry.get("most_recent_onsite_deliverables") or ev.get("onsite_deliverables", "")

    for name, s, e, geo, tier in phase_defs:
        rows.append({
            "campaign_name":       name,
            "canonical_key":       "bts",
            "tier":                tier,
            "geography":           geo,
            "start_date":          s.isoformat(),
            "end_date":            e.isoformat(),
            "duration_days":       calc_duration(s.isoformat(), e.isoformat()),
            "promos_coupons":      "TBC",
            "category_focus":      cat,
            "onsite_deliverables": deliv,
            "wireframes":          "TBD",
            "event_type":          "seasonal",
            "source":              "both" if kb_entry else "config",
            "needs_review":        True,
            "raw_notes":           "Dates estimated from 2025 pattern (offset +1yr). Confirm phasing.",
        })
    return rows


def build_trend_edit_rows(cfg, year, month, kb):
    """Generate trend-edit suggestion rows (all needs_review=TRUE)."""
    if 4 <= month <= 8:
        pool   = cfg.get("trend_edit_pool", {}).get("spring_summer", [])
        season = "spring_summer"
    else:
        pool   = cfg.get("trend_edit_pool", {}).get("autumn_winter", [])
        season = "autumn_winter"

    rows = []
    for edit in pool:
        rows.append({
            "campaign_name":       edit["name"],
            "canonical_key":       season,
            "tier":                "Tier 2",
            "geography":           "KSA + UAE",
            "start_date":          "TBD",
            "end_date":            "TBD",
            "duration_days":       7,
            "promos_coupons":      "NA",
            "category_focus":      edit.get("focus", ""),
            "onsite_deliverables": "Hero Banner",
            "wireframes":          "TBD",
            "event_type":          "trend_edit",
            "source":              "config",
            "needs_review":        True,
            "raw_notes":           "Trend edit suggestion — specialist to select, sequence, and rewrite.",
        })
    return rows


# ── XLSX writer ─────────────────────────────────────────────────────────────────
def write_xlsx(df, filepath, month, year):
    """Write the draft calendar DataFrame to an xlsx matching the june2026 style."""
    import calendar as cal_mod
    month_name = cal_mod.month_name[month]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{month_name[:3]} {year}"

    # ── Title rows (rows 1-3) ──
    ws.merge_cells("A1:U1")
    title_cell = ws["A1"]
    title_cell.value = f"Draft Campaign Calendar — {month_name} {year}"
    title_cell.font  = Font(name="Calibri", bold=True, size=14, color="FF00FF00")
    title_cell.fill  = PatternFill("solid", fgColor="FF000000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:U2")
    sub_cell = ws["A2"]
    sub_cell.value = (
        f"AUTO-GENERATED by Layer 1 Engine  |  "
        f"Config version: 1  |  "
        f"DRAFT — review all rows before use"
    )
    sub_cell.font      = Font(name="Calibri", italic=True, size=9, color="FFAAAAAA")
    sub_cell.fill      = PatternFill("solid", fgColor="FF1A1A1A")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 14

    ws.row_dimensions[3].height = 6   # spacer

    # ── Header row (row 4) ──
    hdr_row = 4
    ws.row_dimensions[hdr_row].height = 36
    for col_idx, hdr in enumerate(DISPLAY_HEADERS, start=1):
        cell = ws.cell(row=hdr_row, column=col_idx, value=hdr)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border    = CELL_BORDER

    # ── Data rows ──
    data_start_row = 5
    for row_idx, record in enumerate(df.to_dict("records"), start=data_start_row):
        ws.row_dimensions[row_idx].height = 50
        for col_idx, col_name in enumerate(SCHEMA_COLS, start=1):
            val    = record.get(col_name, "")
            # Format value for display
            if col_name in LAYER2_COLS:
                val = LAYER2_PLACEHOLDER if not val or val == "" else val
            elif col_name == "needs_review":
                val = "TRUE" if str(val).upper() in ("TRUE", "1", "YES") else "FALSE"
            elif val is None or str(val) in ("nan", "None", ""):
                val = ""

            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else "")
            cell.font      = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border    = CELL_BORDER

            # Light highlight for needs_review rows
            if str(record.get("needs_review", "")).upper() in ("TRUE", "1"):
                cell.fill = PatternFill("solid", fgColor="FFFFF8E1")   # pale amber

    # ── Column widths ──
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Freeze panes below header ──
    ws.freeze_panes = ws.cell(row=data_start_row, column=1)

    wb.save(filepath)
    log.info(f"Wrote {filepath}")


# ── Notes writer ────────────────────────────────────────────────────────────────
def write_notes(rows, cfg, kb, history_df, year, month, filepath):
    """Write draft_notes.md."""
    import calendar as cal_mod

    lines = [
        f"# Draft Campaign Calendar — {cal_mod.month_name[month]} {year}",
        f"*Generated by Layer 1 Engine (config schema_version=1)*",
        "",
        "---",
        "",
        "## 1. Events resolved for this month",
        "",
    ]

    for r in rows:
        nr = "⚠️ needs_review" if str(r.get("needs_review", "")).upper() == "TRUE" else "✓"
        lines.append(
            f"- **{r['campaign_name']}** "
            f"({r['event_type']}) "
            f"{r['start_date']} → {r['end_date']} "
            f"| source={r['source']} | {nr}"
        )

    # Rows that need review + reason
    flagged = [r for r in rows if str(r.get("needs_review", "")).upper() == "TRUE"]
    lines += [
        "",
        "---",
        "",
        f"## 2. Rows flagged `needs_review=TRUE` ({len(flagged)} of {len(rows)})",
        "",
    ]
    for r in flagged:
        note = r.get("raw_notes", "")
        lines.append(f"- **{r['campaign_name']}**: {note or 'Best-guess values used.'}")

    # Config events with no history
    config_only = [r for r in rows if r.get("source") == "config"]
    lines += [
        "",
        "---",
        "",
        f"## 3. Config events with no historical precedent ({len(config_only)} rows)",
        "*(These rows are based purely on the config; no past Namshi data to validate against)*",
        "",
    ]
    for r in config_only:
        lines.append(f"- {r['campaign_name']} ({r['canonical_key']})")

    # History rows with no matching config event
    all_ck_in_config = set()
    for section in ("fixed_date_events", "seasonal_windows", "monthly_anchors",
                    "lunar_events", "sports_events"):
        for ev in cfg.get(section, []):
            all_ck_in_config.add(ev["canonical_key"])

    if history_df is not None and not history_df.empty:
        hist_keys = set(history_df["canonical_key"].dropna().unique())
        orphan_keys = hist_keys - all_ck_in_config - {"unknown", ""}
        lines += [
            "",
            "---",
            "",
            f"## 4. History campaigns with no matching config event ({len(orphan_keys)} keys)",
            "*(Possible gaps — may need a config entry if these recur regularly)*",
            "",
        ]
        for ck in sorted(orphan_keys):
            count = len(history_df[history_df["canonical_key"] == ck])
            lines.append(f"- `{ck}` (appears {count}x in history)")

    # Assumptions
    lines += [
        "",
        "---",
        "",
        "## 5. Assumptions & notes",
        "",
        "- **Dates estimated from 2025 history offset +1 year** — verify before sharing.",
        "- **Lunar dates** (Ramadan, Eid) are config-supplied estimates — must be confirmed.",
        "- **Sports dates** (World Cup, F1) are config-supplied — verify official fixtures.",
        "- **BTS phases** are based on 2025 KSA/UAE split pattern; confirm geo/timing.",
        "- **Trend edits** are suggestions from the pool; specialist should select, sequence, and rewrite.",
        "- **Beauty Week** needs confirmation it is active this month.",
        "- **Layer 2 fields** (trending_keywords, stock_flag, etc.) are all blank placeholders.",
        "- **This draft is never auto-published.** All edits happen in the XLSX before any use.",
    ]

    filepath.write_text("\n".join(lines), encoding="utf-8")
    log.info(f"Wrote {filepath}")


# ── Main ────────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Generate draft campaign calendar")
    parser.add_argument("--month", type=int, required=True, help="Target month (1-12)")
    parser.add_argument("--year",  type=int, required=True, help="Target year (e.g. 2026)")
    args = parser.parse_args()

    month, year = args.month, args.year
    if not (1 <= month <= 12):
        log.error("Month must be 1-12.")
        sys.exit(1)

    log.info(f"Generating draft for {year}-{month:02d}")

    # ── Load inputs ──
    cfg        = load_config()
    history_df = None
    if CSV_FILE.exists():
        history_df = pd.read_csv(CSV_FILE, dtype=str, keep_default_na=False)
        kb = build_kb(history_df)
        log.info(f"Loaded history: {len(history_df)} rows, {len(kb)} canonical keys in KB")
    else:
        log.warning("standardized_history.csv not found — run standardize.py first. Falling back to config only.")
        kb = {}

    # ── Step 3: Resolve events ──
    resolved = resolve_events(cfg, year, month)
    log.info(f"Resolved {len(resolved)} events from config for {year}-{month:02d}")

    # ── Step 4: Generate draft rows ──
    rows = []
    seen_ck = set()

    for ev in resolved:
        ck = ev["canonical_key"]

        # BTS gets special multi-phase handling
        if ck == "bts":
            if "bts" not in seen_ck:
                rows.extend(build_bts_rows(ev, kb, year, month))
                seen_ck.add("bts")
            continue

        # Skip duplicate canonical keys (e.g. EOSS appears in two seasonal windows)
        # but keep them if they have different names (Summer vs Winter EOSS)
        dedup_key = f"{ck}__{ev['name']}"
        if dedup_key in seen_ck:
            continue
        seen_ck.add(dedup_key)

        rows.append(build_row(ev, kb, year, month))

    # Trend edits (added last)
    rows.extend(build_trend_edit_rows(cfg, year, month, kb))

    log.info(f"Generated {len(rows)} draft rows")

    # ── Layer 2 enrichment (no-op stub) ──
    draft_df = pd.DataFrame(rows)
    draft_df = enrich_with_commercial_signals(draft_df)

    # Ensure column order
    for col in SCHEMA_COLS:
        if col not in draft_df.columns:
            draft_df[col] = ""

    draft_df = draft_df[SCHEMA_COLS]

    # ── Step 5: Emit outputs ──
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path  = OUT_DIR / f"draft_calendar_{month}_{year}.xlsx"
    notes_path = OUT_DIR / "draft_notes.md"

    write_xlsx(draft_df, xlsx_path, month, year)
    write_notes(rows, cfg, kb, history_df, year, month, notes_path)

    print(f"\nDone.")
    print(f"  XLSX  : {xlsx_path}")
    print(f"  Notes : {notes_path}")
    print(f"  Rows  : {len(draft_df)}  (needs_review: {draft_df['needs_review'].eq(True).sum()})")


if __name__ == "__main__":
    main()
