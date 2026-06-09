"""
standardize.py — Step 1: Normalize all history sheets → standardized_history.csv

Usage:
    python src/standardize.py

Reads every sheet in Master Campaign Calendr.xlsx, maps each to the Standard Schema,
and writes data/standardized_history.csv. Also prints a parse log at the end.
"""

import re
import logging
from datetime import date, datetime, timedelta
from pathlib import Path

import pandas as pd
import openpyxl
import yaml
from dateutil import parser as duparser

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE     = Path(__file__).resolve().parent.parent
RAW_FILE = BASE / "data" / "raw" / "Master Campaign Calendr.xlsx"
OUT_CSV  = BASE / "data" / "standardized_history.csv"
CFG_FILE = BASE / "config" / "calendar_config.yaml"

logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────
SKIP_SHEETS  = {"april 2025"}
EXCEL_EPOCH  = date(1899, 12, 30)
MIN_HDR_HITS = 3   # minimum keyword matches to call a row a header

STD_COLS = [
    "campaign_name", "canonical_key", "tier", "geography",
    "start_date", "end_date", "duration_days",
    "promos_coupons", "category_focus", "onsite_deliverables",
    "wireframes", "event_type", "source", "needs_review", "raw_notes",
    "sheet_name", "table_type",
]

HEADER_KW = {
    "campaign", "campaigns", "central campaign",
    "geography", "tier",
    "start date", "tentative start date",
    "end date", "tentative end date",
    "promos", "category", "onsite", "wireframe", "deliverable",
}

# Raw header cell (lower) → standard field.
# None  = skip column entirely.
# _xxx  = extra promo field (concatenated into promos_coupons).
COL_MAP = {
    "central campaign":                  "campaign_name",
    "campaigns":                         "campaign_name",
    "trends":                            "campaign_name",
    "theme":                             "campaign_name",
    "tier":                              "tier",
    "geography":                         "geography",
    "country":                           "geography",
    "tentative start date":              "start_date",
    "start date":                        "start_date",
    "tentative end date":                "end_date",
    "end date":                          "end_date",
    "promos + coupons":                  "promos_coupons",
    "promos":                            "promos_coupons",
    "beauty callout":                    "_beauty_callout",
    "cashback coupons":                  "_cashback_coupons",
    "coupon change":                     "_coupon_change",
    "deal codes":                        "_deal_codes",
    "category focus":                    "category_focus",
    "merch focus":                       "category_focus",
    "onsite plan/ deliverables (tbc)":   "onsite_deliverables",
    "onsite deliverables":               "onsite_deliverables",
    "wireframes":                        "wireframes",
    "wireframe":                         "wireframes",
    # Ignored columns
    "crm plan/ deliverables":            None,
    "dm & social plan/ deliverables":    None,
    "bp plan/ deliverables":             None,
    "central poc":                       None,
    "comments/ additional ref":          None,
    "time":                              None,
    "objective/key messaging":           None,
}

MONTH_NAMES = {
    "jan": 1, "january": 1, "feb": 2, "february": 2,
    "mar": 3, "march": 3,  "apr": 4, "april": 4,
    "may": 5, "jun": 6, "june": 6,  "jul": 7, "july": 7,
    "aug": 8, "august": 8, "sep": 9, "september": 9,
    "oct": 10, "october": 10, "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}


# ── Config loading ──────────────────────────────────────────────────────────────
def load_config():
    with open(CFG_FILE, encoding="utf-8") as f:
        return yaml.safe_load(f)


def build_canonical_map(cfg):
    """Returns dict: campaign name (lower, stripped) → (canonical_key, event_type)."""
    m = {}
    for section in ("fixed_date_events", "seasonal_windows", "monthly_anchors",
                    "lunar_events", "sports_events"):
        for ev in cfg.get(section, []):
            m[ev["name"].lower().strip()] = (ev["canonical_key"], ev["event_type"])

    # Config-supplied name variants
    for raw_name, ck in cfg.get("campaign_name_map", {}).items():
        # Derive event_type from canonical_key lookup
        et = _event_type_for_key(ck, cfg)
        m[raw_name.lower().strip()] = (ck, et)

    # Trend pool names
    for season, edits in cfg.get("trend_edit_pool", {}).items():
        for edit in edits:
            m[edit["name"].lower().strip()] = (season, "trend_edit")

    return m


def _event_type_for_key(ck, cfg):
    """Look up the event_type for a canonical_key in config."""
    for section in ("fixed_date_events", "seasonal_windows", "monthly_anchors",
                    "lunar_events", "sports_events"):
        for ev in cfg.get(section, []):
            if ev["canonical_key"] == ck:
                return ev["event_type"]
    return ""


# ── Sheet date helpers ──────────────────────────────────────────────────────────
def parse_sheet_sort_key(sheet_name):
    """Return (year, month) integer tuple for sorting sheets chronologically."""
    name = sheet_name.lower().strip()
    year_match = re.search(r"(20\d{2})", name)
    if not year_match:
        return (0, 0)
    year = int(year_match.group(1))
    name_no_yr = name.replace(year_match.group(1), "").strip()
    for mname, mnum in MONTH_NAMES.items():
        if mname in name_no_yr:
            return (year, mnum)
    return (year, 0)


# ── Header detection ────────────────────────────────────────────────────────────
def header_score(row):
    """Count how many cells in the row match a header keyword."""
    score = 0
    for cell in row:
        if cell is None:
            continue
        cell_lower = str(cell).lower().strip()
        if any(kw in cell_lower for kw in HEADER_KW):
            score += 1
    return score


def is_trend_header(row):
    """True if the first non-None cell is exactly 'trends' or 'trend'."""
    for cell in row:
        if cell is None:
            continue
        return str(cell).strip().lower() in ("trends", "trend")
    return False


def is_empty_row(row):
    return all(c is None or str(c).strip() == "" for c in row)


# ── Column mapping ──────────────────────────────────────────────────────────────
def map_header(header_row):
    """Map each column index to a field name (or None to skip)."""
    fields = []
    for cell in header_row:
        if cell is None:
            fields.append(None)
            continue
        key = str(cell).strip().lower()
        field = COL_MAP.get(key)
        if field is None and key not in COL_MAP:
            # Partial-match fallback
            matched = None
            for pattern, fname in COL_MAP.items():
                if pattern in key:
                    matched = fname
                    break
            fields.append(matched)
        else:
            fields.append(field)
    return fields


# ── Date conversion ─────────────────────────────────────────────────────────────
def parse_cell_date(value):
    """
    Convert cell value to (iso_str_or_TBD, raw_str_if_flagged, needs_review_flag).
    """
    if value is None:
        return "TBD", None, False

    # Already a Python date/datetime
    if isinstance(value, datetime):
        return value.date().isoformat(), None, False
    if isinstance(value, date):
        return value.isoformat(), None, False

    # Excel serial number stored as float/int
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        try:
            converted = EXCEL_EPOCH + timedelta(days=int(value))
            return converted.isoformat(), str(value), False
        except Exception:
            pass

    # String
    if isinstance(value, str):
        s = value.strip()
        if s in ("-", "", "NA", "na", "N/A"):
            return "TBD", None, False
        if s.upper() in ("TBD", "TBC"):
            return "TBD", s, False
        # Try dateutil fuzzy parse
        try:
            parsed = duparser.parse(s, fuzzy=True, dayfirst=False)
            return parsed.date().isoformat(), s, True   # raw preserved, review flagged
        except Exception:
            return "TBD", s, True   # unparseable → TBD + flag

    return "TBD", str(value), True


# ── Normalizers ─────────────────────────────────────────────────────────────────
def normalize_geo(value, geo_norms):
    if value is None:
        return None, None
    raw = str(value).strip()
    clean = raw.lower()
    # Exact match in config table
    norm = geo_norms.get(clean) or geo_norms.get(clean.replace(" ", ""))
    if norm:
        return norm, None
    # Fallback patterns
    if "ksa" in clean and "uae" in clean:
        return "KSA + UAE", raw
    if "ksa" in clean:
        return "KSA", raw
    if "uae" in clean or ("ae" in clean and len(clean) <= 5):
        return "UAE", raw
    if clean in ("all", "all "):
        return "All", None
    if "gcc" in clean:
        return "UAE", raw
    return None, raw


def normalize_tier(value, tier_norms):
    if value is None:
        return "-", None
    raw = str(value).strip()
    clean = raw.lower()
    norm = tier_norms.get(clean)
    if norm is not None:
        return norm, None
    # Complex multi-tier: extract highest tier and preserve raw
    if "tier 1" in clean or "\nt1" in clean or "t1 " in clean:
        return "Tier 1", raw
    if "tier 2" in clean or "t2" in clean:
        return "Tier 2", raw
    if "tier 3" in clean or "t3" in clean:
        return "Tier 3", raw
    return "-", raw


# ── Row processing ──────────────────────────────────────────────────────────────
def process_row(values, fields, geo_norms, tier_norms):
    """
    Build a partial record dict from a data row.
    Returns (record_dict, needs_review_flag).
    """
    rec = {
        "campaign_name": None, "tier": "-", "geography": None,
        "start_date": None, "end_date": None,
        "promos_coupons": None, "category_focus": None,
        "onsite_deliverables": None, "wireframes": None,
    }
    promo_extras = {}
    raw_parts    = []
    needs_review = False

    for field, val in zip(fields, values):
        if field is None:
            continue

        if field == "campaign_name":
            if val is not None:
                rec["campaign_name"] = str(val).strip()

        elif field == "tier":
            norm, raw = normalize_tier(val, tier_norms)
            rec["tier"] = norm
            if raw:
                raw_parts.append(f"tier_raw: {raw}")
                needs_review = True

        elif field == "geography":
            norm, raw = normalize_geo(val, geo_norms)
            rec["geography"] = norm
            if raw and norm is None:
                raw_parts.append(f"geo_raw: {raw}")
                needs_review = True

        elif field == "start_date":
            iso, raw, flagged = parse_cell_date(val)
            rec["start_date"] = iso
            if raw:
                raw_parts.append(f"start_raw: {raw}")
            if flagged:
                needs_review = True

        elif field == "end_date":
            iso, raw, flagged = parse_cell_date(val)
            rec["end_date"] = iso
            if raw:
                raw_parts.append(f"end_raw: {raw}")
            if flagged:
                needs_review = True

        elif field == "promos_coupons":
            if val is not None and str(val).strip() not in ("", "-"):
                rec["promos_coupons"] = str(val).strip()

        elif field.startswith("_"):
            key = field[1:]
            if val is not None and str(val).strip() not in ("", "-", "NA", "na"):
                promo_extras[key] = str(val).strip()

        elif field in ("category_focus", "onsite_deliverables", "wireframes"):
            if val is not None and str(val).strip() not in ("", "-"):
                rec[field] = str(val).strip()

    # Merge extra promo columns
    if promo_extras:
        extra_str = " | ".join(f"{k}: {v}" for k, v in promo_extras.items())
        if rec["promos_coupons"]:
            rec["promos_coupons"] = rec["promos_coupons"] + " | " + extra_str
        else:
            rec["promos_coupons"] = extra_str

    rec["raw_notes"] = " | ".join(raw_parts) if raw_parts else None
    return rec, needs_review


# ── Sheet parsing ───────────────────────────────────────────────────────────────
def parse_sheet(ws, sheet_name, canonical_map, geo_norms, tier_norms, parse_log):
    """
    Extract all rows from a sheet. Returns list of record dicts (partial).
    Appends messages to parse_log.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    records  = []

    primary_header = None
    primary_fields = None
    trend_header   = None
    trend_fields   = None
    current_table  = None   # "primary" | "trend"

    for row_idx, row in enumerate(all_rows, start=1):
        if is_empty_row(row):
            continue

        # ── Determine which table context we're in ──
        if is_trend_header(row) and header_score(row) >= MIN_HDR_HITS and primary_header is not None:
            # Switch to trend table
            trend_header  = row
            trend_fields  = map_header(row)
            current_table = "trend"
            log.info(f"  [{sheet_name}] Trend table detected at row {row_idx}")
            continue

        if header_score(row) >= MIN_HDR_HITS:
            if primary_header is None:
                primary_header = row
                primary_fields = map_header(row)
                current_table  = "primary"
                log.info(f"  [{sheet_name}] Primary header at row {row_idx}, cols={[c for c in row if c]}")
                continue
            elif current_table == "trend" and trend_header is None:
                trend_header  = row
                trend_fields  = map_header(row)
                continue
            # else: ignore duplicate headers
            continue

        # ── Data row ──
        if primary_header is None:
            continue  # Haven't found a header yet

        active_fields = trend_fields if current_table == "trend" else primary_fields
        table_type    = current_table

        rec, nr = process_row(row, active_fields, geo_norms, tier_norms)

        if not rec.get("campaign_name"):
            parse_log.append(f"[{sheet_name}] Row {row_idx}: no campaign_name, skipped: {row[:4]}")
            continue

        # Resolve canonical key + event_type
        ck_key = rec["campaign_name"].lower().strip()
        canon_match = canonical_map.get(ck_key)
        if not canon_match:
            # Try partial
            for name_pat, (ck, et) in canonical_map.items():
                if name_pat in ck_key or ck_key in name_pat:
                    canon_match = (ck, et)
                    break
        if canon_match:
            rec["canonical_key"] = canon_match[0]
            rec["event_type"]    = canon_match[1]
        else:
            rec["canonical_key"] = "unknown"
            rec["event_type"]    = ""
            parse_log.append(f"[{sheet_name}] Row {row_idx}: unmatched campaign '{rec['campaign_name']}'")
            nr = True

        # Duration
        rec["duration_days"] = _calc_duration(rec["start_date"], rec["end_date"])

        rec["source"]       = "history"
        rec["needs_review"] = nr
        rec["sheet_name"]   = sheet_name
        rec["table_type"]   = table_type
        # sheet_sort_key stored as tuple, we'll convert to yyyymm int for CSV
        rec["sheet_sort_key"] = parse_sheet_sort_key(sheet_name)

        records.append(rec)

    return records


def _calc_duration(start, end):
    if start and end and start != "TBD" and end != "TBD":
        try:
            return (date.fromisoformat(end) - date.fromisoformat(start)).days + 1
        except Exception:
            pass
    return None


# ── Main ────────────────────────────────────────────────────────────────────────
def main():
    cfg          = load_config()
    canonical_map = build_canonical_map(cfg)
    geo_norms    = {k.lower(): v for k, v in cfg.get("geography_normalization", {}).items()}
    tier_norms   = {k.lower(): v for k, v in cfg.get("tier_normalization", {}).items()}

    wb        = openpyxl.load_workbook(RAW_FILE, data_only=True)
    parse_log = []
    all_recs  = []

    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() in SKIP_SHEETS:
            log.info(f"Skipping sheet: {sheet_name}")
            continue
        log.info(f"Processing: {sheet_name}")
        ws   = wb[sheet_name]
        recs = parse_sheet(ws, sheet_name, canonical_map, geo_norms, tier_norms, parse_log)
        log.info(f"  → {len(recs)} rows")
        all_recs.extend(recs)

    if not all_recs:
        log.error("No records extracted. Check file path and sheet format.")
        return

    df = pd.DataFrame(all_recs)

    # Sort by sheet date descending (most recent first), then by row order
    df["_sort"] = df["sheet_sort_key"].apply(lambda t: t[0] * 100 + t[1])
    df = df.sort_values("_sort", ascending=False).reset_index(drop=True)
    df.drop(columns=["_sort", "sheet_sort_key"], inplace=True)

    # Ensure all standard columns present
    for col in STD_COLS:
        if col not in df.columns:
            df[col] = None

    df = df[STD_COLS]

    OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUT_CSV, index=False, encoding="utf-8-sig")
    log.info(f"\nWrote {len(df)} rows → {OUT_CSV}")

    # ── Parse report ──
    print("\n" + "=" * 60)
    print("PARSE LOG")
    print("=" * 60)
    if parse_log:
        for msg in parse_log:
            print(" •", msg)
    else:
        print("  No issues.")

    # Summary stats
    print(f"\nTotal rows     : {len(df)}")
    print(f"Sheets parsed  : {df['sheet_name'].nunique()}")
    print(f"Unique keys    : {df['canonical_key'].nunique()}")
    print(f"needs_review   : {df['needs_review'].sum()}")
    print(f"Trend rows     : {(df['table_type'] == 'trend').sum()}")
    print(f"Unknown keys   : {(df['canonical_key'] == 'unknown').sum()}")
    print("=" * 60)


if __name__ == "__main__":
    main()
